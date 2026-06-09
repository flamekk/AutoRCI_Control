from __future__ import annotations

import math
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from src.audit import enrich_report_with_audits
    from src.action_plan import (
        REFERENCE_SUGGESTION_COLUMNS,
        build_reference_suggestions,
        sort_action_plan_records,
    )
    from src.missing_rci import (
        ABSENT_RCI_COLUMNS,
        ABSENT_RCI_HEADERS,
        NO_MISSING_RCI_MESSAGE,
        build_missing_rci_records,
        missing_rci_summary,
    )
except ModuleNotFoundError:  # pragma: no cover - used when running python src/main.py.
    from audit import enrich_report_with_audits
    from action_plan import (
        REFERENCE_SUGGESTION_COLUMNS,
        build_reference_suggestions,
        sort_action_plan_records,
    )
    from missing_rci import (
        ABSENT_RCI_COLUMNS,
        ABSENT_RCI_HEADERS,
        NO_MISSING_RCI_MESSAGE,
        build_missing_rci_records,
        missing_rci_summary,
    )

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ModuleNotFoundError:  # pragma: no cover - exercised only before dependency install.
    Workbook = None  # type: ignore[assignment]
    Alignment = Border = Font = PatternFill = Side = None  # type: ignore[assignment]
    Table = TableStyleInfo = None  # type: ignore[assignment]
    get_column_letter = None  # type: ignore[assignment]


RECONCILIATION_COLUMNS = [
    "invoice_number",
    "document_type",
    "customer_code",
    "customer_name",
    "is_rci_covered",
    "amount_erp",
    "amount_rci",
    "amount_pdf",
    "amount_gap",
    "montant_impacte",
    "erp_date",
    "rci_date",
    "pdf_invoice_date",
    "due_date",
    "origin",
    "status",
    "priority",
    "severity",
    "included_in_corrective_batch",
    "action_recommandee",
    "source_erp",
    "source_rci",
    "source_pdf",
    "commentaire_audit",
]

RECONCILIATION_HEADERS = {
    "invoice_number": "N° facture",
    "document_type": "Type document",
    "customer_code": "Code concessionnaire",
    "customer_name": "Concessionnaire",
    "is_rci_covered": "Couvert RCI",
    "amount_erp": "Montant ERP",
    "amount_rci": "Montant RCI",
    "amount_pdf": "Montant PDF",
    "amount_gap": "Ecart montant",
    "montant_impacte": "Montant impacte",
    "erp_date": "Date ERP",
    "rci_date": "Date RCI",
    "pdf_invoice_date": "Date facture PDF",
    "due_date": "Date échéance",
    "origin": "Origine",
    "status": "Statut",
    "priority": "Priorité",
    "severity": "Sévérité",
    "included_in_corrective_batch": "Inclus batch correctif",
    "action_recommandee": "Action recommandée",
    "source_erp": "Source ERP",
    "source_rci": "Source RCI",
    "source_pdf": "Source PDF",
    "commentaire_audit": "Commentaire audit",
}

STATUS_FILLS = {
    "OK": "C6EFCE",
    "MANQUANTE_RCI": "FCE4D6",
    "ANOMALIE_MONTANT": "FFC7CE",
    "ANOMALIE_DATE": "FFC7CE",
    "DOUBLON": "E4DFEC",
    "RCI_SEULEMENT": "D9EAF7",
    "RCI_HORS_PERIODE": "DDEBF7",
    "HORS_SCOPE_RCI": "E7E6E6",
}

SEVERITY_FILLS = {
    "CRITIQUE": "FFC7CE",
    "ELEVEE": "FCE4D6",
    "MOYENNE": "FFF2CC",
    "A_VERIFIER": "E7E6E6",
    "INFORMATION": "DDEBF7",
    "OK": "C6EFCE",
}

SEVERITY_FONT_COLORS = {
    "CRITIQUE": "9C0006",
    "ELEVEE": "9C5700",
    "MOYENNE": "7F6000",
    "A_VERIFIER": "666666",
    "INFORMATION": "1F4E78",
    "OK": "006100",
}

STATUS_FONT_COLORS = {
    "OK": "006100",
    "MANQUANTE_RCI": "9C5700",
    "ANOMALIE_MONTANT": "9C0006",
    "ANOMALIE_DATE": "9C0006",
    "DOUBLON": "5F497A",
    "RCI_SEULEMENT": "1F4E78",
    "RCI_HORS_PERIODE": "1F4E78",
    "HORS_SCOPE_RCI": "666666",
}

MONEY_COLUMNS = {"amount_erp", "amount_rci", "amount_pdf", "amount_gap", "montant_impacte"}
SUMMARY_MONEY_LABELS = {
    "Montant total contrôlé",
    "Montant impacté total",
    "Montant manquant RCI",
    "Montant total absent RCI",
    "Montant total inclus batch correctif",
}
SUMMARY_PERCENT_LABELS = {
    "Taux de rapprochement",
    "Pourcentage hors scope RCI",
}
TABLE_STYLE = "TableStyleMedium2"
MAX_TABLE_ROWS = 5000
MAX_AUTOFIT_ROWS = 1000

AUDIT_DATES_COLUMNS = [
    "erp_date",
    "nombre_total_factures_erp",
    "nombre_ok",
    "nombre_manquante_rci",
    "nombre_hors_scope_rci",
    "montant_manquant_rci",
    "taux_rapprochement_date",
    "rci_count",
    "rci_hors_periode_count",
]

AUDIT_MISSING_RCI_COLUMNS = [
    "invoice_number",
    "erp_date",
    "customer_name",
    "customer_code",
    "amount_erp",
    "is_rci_covered",
    "status",
    "montant_impacte",
    "source_erp",
    "date_in_filter",
    "commentaire_audit",
]

RCI_OUT_OF_PERIOD_COLUMNS = [
    "invoice_number",
    "document_type",
    "rci_date",
    "pdf_invoice_date",
    "amount_rci",
    "amount_pdf",
    "source_rci",
    "source_pdf",
    "status",
    "montant_impacte",
    "commentaire_audit",
]

AUDIT_OUT_OF_SCOPE_COLUMNS = [
    "invoice_number",
    "erp_date",
    "customer_name",
    "normalized_customer_name",
    "amount_erp",
    "status",
    "closest_reference_name",
    "closest_reference_similarity",
    "commentaire_audit",
]

ACTION_PLAN_COLUMNS = [
    "severity",
    "invoice_number",
    "erp_date",
    "customer_name",
    "amount_erp",
    "amount_rci",
    "amount_pdf",
    "montant_impacte",
    "status",
    "action_recommandee",
    "source_erp",
    "source_rci",
    "source_pdf",
]

BATCH_CORRECTIF_COLUMNS = [
    "invoice_number",
    "erp_date",
    "customer_name",
    "amount_erp",
    "montant_impacte",
    "severity",
    "status",
    "action_recommandee",
]


def write_excel_report(report: dict[str, Any], output_dir: Path, run_id: str) -> Path:
    workbook_class = _require_openpyxl()
    if not report.get("audits"):
        report = enrich_report_with_audits(report, log=False)
    output_dir.mkdir(parents=True, exist_ok=True)
    generated_at = _report_datetime(report, run_id)
    workbook_path = output_dir / f"Rapport_Reconciliation_RCI_{generated_at:%Y-%m-%d_%H%M}.xlsx"

    workbook = workbook_class()
    workbook.remove(workbook.active)

    reconciliation_rows = list(report.get("reconciliation", []))
    missing_rci_rows = build_missing_rci_records(reconciliation_rows)
    report["missing_rci_records"] = missing_rci_rows
    report.setdefault("summary", {}).update(missing_rci_summary(reconciliation_rows))
    action_plan_rows = [
        row for row in sort_action_plan_records(reconciliation_rows)
        if row.get("severity") != "INFORMATION"
    ]
    audits = report.get("audits", {})

    _write_rows_sheet(workbook, "Synthèse", _summary_rows(report, generated_at), table_name="TableSynthese")
    _write_missing_rci_sheet(workbook, missing_rci_rows, "TableAbsentsRCI")
    _write_reconciliation_sheet(workbook, "Détail rapprochement", reconciliation_rows, "TableDetail")
    _write_action_plan_sheet(
        workbook,
        action_plan_rows,
        "TablePlanAction",
    )
    _write_batch_correctif_sheet(
        workbook,
        report.get("corrective_batch", {}),
        "TableBatchCorrectif",
    )
    _write_reconciliation_sheet(
        workbook,
        "Hors périmètre RCI",
        _filter_status(reconciliation_rows, {"HORS_SCOPE_RCI"}),
        "TableHorsPerimetreRCI",
    )
    _write_rows_sheet(
        workbook,
        "RCI PDF hors période",
        _dict_rows(_filter_status(reconciliation_rows, {"RCI_HORS_PERIODE"}), RCI_OUT_OF_PERIOD_COLUMNS),
        table_name="TableRCIPDFHorsPeriode",
    )
    _write_rows_sheet(
        workbook,
        "Qualité référentiel RCI",
        _dict_rows(
            report.get("reference_quality") or build_reference_suggestions(reconciliation_rows),
            REFERENCE_SUGGESTION_COLUMNS,
        ),
        table_name="TableQualiteReferentielRCI",
    )
    _write_rows_sheet(
        workbook,
        "Synthèse par concessionnaire",
        _dealer_summary_rows(reconciliation_rows),
        table_name="TableSyntheseConcessionnaire",
    )
    _write_audit_sheet(workbook, audits, "TableAudit")

    workbook.save(workbook_path)
    return workbook_path


def _summary_rows(report: dict[str, Any], generated_at: datetime) -> list[list[Any]]:
    summary = report.get("summary", {})
    reconciliation = report.get("reconciliation", [])
    status_counts = _status_counts(reconciliation)
    ok_count = status_counts.get("OK", int(summary.get("matched_invoices", 0) or 0))
    missing_count = status_counts.get("MANQUANTE_RCI", int(summary.get("unmatched_erp", 0) or 0))
    amount_anomalies = status_counts.get("ANOMALIE_MONTANT", int(summary.get("amount_anomalies", 0) or 0))
    date_anomalies = status_counts.get("ANOMALIE_DATE", int(summary.get("date_anomalies", 0) or 0))
    duplicates = status_counts.get("DOUBLON", int(summary.get("duplicates", 0) or 0))
    rci_only = status_counts.get("RCI_SEULEMENT", int(summary.get("unmatched_rci", 0) or 0))
    rci_out_of_period = status_counts.get("RCI_HORS_PERIODE", int(summary.get("rci_out_of_period", 0) or 0))
    out_of_scope = status_counts.get("HORS_SCOPE_RCI", int(summary.get("out_of_scope_rci", 0) or 0))
    analyzed_total = int(summary.get("erp_analyzed_invoices", summary.get("reconciled_invoices", 0)) or 0)
    if analyzed_total == 0:
        analyzed_total = ok_count + missing_count + out_of_scope + amount_anomalies + date_anomalies + duplicates
    missing_summary = missing_rci_summary(reconciliation)
    missing_invoice_count = int(summary.get("missing_rci_invoice_count", missing_summary["missing_rci_invoice_count"]) or 0)
    missing_credit_note_count = int(
        summary.get("missing_rci_credit_note_count", missing_summary["missing_rci_credit_note_count"]) or 0
    )
    missing_rci_amount = float(summary.get("missing_rci_total_amount", missing_summary["missing_rci_total_amount"]) or 0)
    erp_matchable = int(summary.get("erp_matchable_invoices", 0) or (
        ok_count + missing_count + amount_anomalies + date_anomalies + duplicates
    ))
    gaps_detected = int(summary.get("gaps_detected", 0) or (
        missing_count + amount_anomalies + date_anomalies + duplicates + rci_only
    ))
    matching_rate = float(summary.get("matching_rate", 0) or (ok_count / erp_matchable if erp_matchable else 0))
    present_in_rci_count = ok_count + amount_anomalies + date_anomalies + duplicates

    return [
        ["Indicateur", "Valeur"],
        ["Résultat du contrôle batch ERP vs batch RCI", ""],
        ["Date de traitement", generated_at.strftime("%Y-%m-%d %H:%M")],
        ["Période de rapprochement", summary.get("reconciliation_period", "filtre desactive")],
        ["Nombre total de factures/avoirs ERP analysés", analyzed_total],
        ["Nombre de factures présentes dans RCI", present_in_rci_count],
        ["Nombre de factures absentes RCI", missing_invoice_count],
        ["Nombre d'avoirs absents RCI", missing_credit_note_count],
        ["Montant total absent RCI", missing_rci_amount],
        ["Nombre d'écarts critiques", int(summary.get("missing_rci_critical_count", missing_summary["missing_rci_critical_count"]) or 0)],
        ["Nombre d'écarts élevés", int(summary.get("missing_rci_high_count", missing_summary["missing_rci_high_count"]) or 0)],
        ["Nombre d'écarts moyens", int(summary.get("missing_rci_medium_count", missing_summary["missing_rci_medium_count"]) or 0)],
        ["Taux de rapprochement", matching_rate],
        ["Informations complémentaires", ""],
        ["Factures dans le périmètre RCI", erp_matchable],
        ["Factures hors périmètre RCI", out_of_scope],
        ["Écarts détectés", gaps_detected],
        ["RCI seulement", rci_only],
        ["RCI hors période", rci_out_of_period],
        ["Total lignes RCI/PDF hors période", int(summary.get("total_rci_pdf_out_of_period", rci_out_of_period) or 0)],
        ["Alerte aucun flux RCI dans la période", _yes_no(summary.get("no_rci_flux_in_period_alert", False))],
        ["Alerte PDF hors période RCI", _yes_no(summary.get("pdf_period_mismatch_alert", False))],
        ["Batch correctif candidat généré", _yes_no(summary.get("corrective_batch_generated", False))],
        ["Nombre factures incluses batch correctif", int(summary.get("corrective_batch_invoice_count", 0) or 0)],
        ["Montant total inclus batch correctif", float(summary.get("corrective_batch_total_amount", 0) or 0)],
    ]


def _write_reconciliation_sheet(
    workbook: Any, sheet_name: str, records: list[dict[str, Any]], table_name: str
) -> None:
    rows = [[RECONCILIATION_HEADERS[column] for column in RECONCILIATION_COLUMNS]]
    for record in records:
        rows.append([_clean_cell_value(record.get(column)) for column in RECONCILIATION_COLUMNS])

    worksheet = _write_rows_sheet(workbook, sheet_name, rows, table_name=table_name)
    status_column = RECONCILIATION_COLUMNS.index("status") + 1
    severity_column = RECONCILIATION_COLUMNS.index("severity") + 1
    for row_index in range(2, worksheet.max_row + 1):
        _apply_status_style(worksheet, row_index, status_column)
        _apply_severity_style(worksheet, row_index, severity_column)
    _format_money_columns(worksheet, RECONCILIATION_COLUMNS)


def _write_missing_rci_sheet(workbook: Any, records: list[dict[str, Any]], table_name: str) -> None:
    if not records:
        worksheet = _write_rows_sheet(
            workbook,
            "Factures et avoirs absents RCI",
            [["Message"], [NO_MISSING_RCI_MESSAGE]],
            table_name=table_name,
        )
        worksheet["A2"].font = Font(bold=True, color="006100")
        return

    rows = [[ABSENT_RCI_HEADERS[column] for column in ABSENT_RCI_COLUMNS]]
    for record in records:
        rows.append([_clean_cell_value(record.get(column)) for column in ABSENT_RCI_COLUMNS])

    worksheet = _write_rows_sheet(workbook, "Factures et avoirs absents RCI", rows, table_name=table_name)
    severity_column = ABSENT_RCI_COLUMNS.index("severity") + 1
    status_column = ABSENT_RCI_COLUMNS.index("status") + 1
    for row_index in range(2, worksheet.max_row + 1):
        _apply_severity_style(worksheet, row_index, severity_column)
        _apply_status_style(worksheet, row_index, status_column)
    _format_money_columns(worksheet, ABSENT_RCI_COLUMNS)


def _write_action_plan_sheet(workbook: Any, records: list[dict[str, Any]], table_name: str) -> None:
    rows = [[RECONCILIATION_HEADERS.get(column, column) for column in ACTION_PLAN_COLUMNS]]
    for record in records:
        rows.append([_clean_cell_value(record.get(column)) for column in ACTION_PLAN_COLUMNS])

    worksheet = _write_rows_sheet(workbook, "Plan action", rows, table_name=table_name)
    severity_column = ACTION_PLAN_COLUMNS.index("severity") + 1
    status_column = ACTION_PLAN_COLUMNS.index("status") + 1
    for row_index in range(2, worksheet.max_row + 1):
        _apply_severity_style(worksheet, row_index, severity_column)
        _apply_status_style(worksheet, row_index, status_column)
    _format_money_columns(worksheet, ACTION_PLAN_COLUMNS)


def _write_batch_correctif_sheet(workbook: Any, batch: dict[str, Any], table_name: str) -> None:
    records = list(batch.get("records", []))
    rows: list[list[Any]] = [
        ["Indicateur", "Valeur"],
        ["Chemin du batch généré", batch.get("batch_path", "")],
        ["Chemin du fichier de contrôle", batch.get("control_path", "")],
        ["Nombre de factures incluses", int(batch.get("invoice_count", 0) or 0)],
        ["Montant total inclus", float(batch.get("total_amount", 0) or 0)],
        [
            "Avertissement",
            batch.get(
                "warning",
                "Fichier candidat à valider par l’équipe facturation avant transmission à RCI.",
            ),
        ],
        [],
        [RECONCILIATION_HEADERS.get(column, column) for column in BATCH_CORRECTIF_COLUMNS],
    ]
    for record in records:
        rows.append([_clean_cell_value(record.get(column)) for column in BATCH_CORRECTIF_COLUMNS])

    worksheet = workbook.create_sheet(title="Batch correctif")
    for row in rows:
        worksheet.append([_clean_cell_value(value) for value in row])

    _style_batch_correctif_sheet(worksheet, table_name, len(records))


def _style_batch_correctif_sheet(worksheet: Any, table_name: str, record_count: int) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index in range(2, 7):
        worksheet.cell(row=row_index, column=1).font = Font(bold=True)
    worksheet.cell(row=5, column=2).number_format = '#,##0.00'
    worksheet.cell(row=6, column=2).font = Font(bold=True, color="9C5700")
    worksheet.cell(row=6, column=2).alignment = Alignment(wrap_text=True)

    detail_header_row = 8
    for cell in worksheet[detail_header_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if record_count:
        ref = f"A{detail_header_row}:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
        table = Table(displayName=_safe_table_name(table_name), ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name=TABLE_STYLE,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
        severity_column = BATCH_CORRECTIF_COLUMNS.index("severity") + 1
        status_column = BATCH_CORRECTIF_COLUMNS.index("status") + 1
        for row_index in range(detail_header_row + 1, worksheet.max_row + 1):
            _apply_severity_style(worksheet, row_index, severity_column)
            _apply_status_style(worksheet, row_index, status_column)
        _format_columns_by_index(worksheet, {4, 5}, '#,##0.00')

    worksheet.freeze_panes = f"A{detail_header_row + 1}"
    _auto_fit_columns(worksheet)


def _write_audit_sheet(workbook: Any, audits: dict[str, Any], table_name: str) -> None:
    worksheet = workbook.create_sheet(title="Audit")
    section_specs = [
        ("Audit dates", AUDIT_DATES_COLUMNS, audits.get("dates", [])),
        ("Audit manquantes RCI", AUDIT_MISSING_RCI_COLUMNS, audits.get("missing_rci", [])),
        ("Audit hors scope RCI", AUDIT_OUT_OF_SCOPE_COLUMNS, audits.get("out_of_scope_rci", [])),
    ]
    for section_name, columns, records in section_specs:
        if worksheet.max_row > 1:
            worksheet.append([])
        section_row = worksheet.max_row + 1
        worksheet.append([section_name])
        worksheet.append(columns)
        for record in records:
            worksheet.append([_clean_cell_value(record.get(column)) for column in columns])
        if not records:
            worksheet.append(["Aucune donnée disponible."])
        worksheet.cell(row=section_row, column=1).font = Font(bold=True, color="1F4E78")

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for row in worksheet.iter_rows():
        first_value = row[0].value
        if first_value in {"Audit dates", "Audit manquantes RCI", "Audit hors scope RCI"}:
            continue
        if any(cell.value in AUDIT_DATES_COLUMNS + AUDIT_MISSING_RCI_COLUMNS + AUDIT_OUT_OF_SCOPE_COLUMNS for cell in row):
            for cell in row:
                if cell.value is None:
                    continue
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    worksheet.freeze_panes = "A3"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{max(worksheet.max_row, 1)}"
    _format_columns_by_index(worksheet, {5, 6, 8}, '#,##0.00')
    _auto_fit_columns(worksheet)


def _dealer_summary_rows(records: list[dict[str, Any]]) -> list[list[Any]]:
    headers = [
        "Code concessionnaire",
        "Concessionnaire",
        "Total factures",
        "OK",
        "MANQUANTE_RCI",
        "ANOMALIE_MONTANT",
        "ANOMALIE_DATE",
        "DOUBLON",
        "RCI_SEULEMENT",
        "RCI_HORS_PERIODE",
        "HORS_SCOPE_RCI",
        "Montant ERP",
        "Montant RCI",
        "Montant PDF",
        "Montant écarts",
        "Montant impacte",
    ]
    groups: dict[tuple[str, str], dict[str, Any]] = {}
    for record in records:
        code = str(record.get("customer_code") or "NON RENSEIGNE")
        name = str(record.get("customer_name") or "NON RENSEIGNE")
        key = (code, name)
        if key not in groups:
            groups[key] = {
                "total": 0,
                "OK": 0,
                "MANQUANTE_RCI": 0,
                "ANOMALIE_MONTANT": 0,
                "ANOMALIE_DATE": 0,
                "DOUBLON": 0,
                "RCI_SEULEMENT": 0,
                "RCI_HORS_PERIODE": 0,
                "HORS_SCOPE_RCI": 0,
                "amount_erp": 0.0,
                "amount_rci": 0.0,
                "amount_pdf": 0.0,
                "amount_gap": 0.0,
                "montant_impacte": 0.0,
            }
        group = groups[key]
        status = str(record.get("status") or "")
        group["total"] += 1
        if status in STATUS_FILLS:
            group[status] += 1
        group["amount_erp"] += abs(_number(record.get("amount_erp")) or 0)
        group["amount_rci"] += abs(_number(record.get("amount_rci")) or 0)
        group["amount_pdf"] += abs(_number(record.get("amount_pdf")) or 0)
        group["amount_gap"] += abs(_number(record.get("amount_gap")) or 0)
        group["montant_impacte"] += abs(_number(record.get("montant_impacte")) or 0)

    rows = [headers]
    for (code, name), values in sorted(groups.items(), key=lambda item: item[0]):
        rows.append(
            [
                code,
                name,
                values["total"],
                values["OK"],
                values["MANQUANTE_RCI"],
                values["ANOMALIE_MONTANT"],
                values["ANOMALIE_DATE"],
                values["DOUBLON"],
                values["RCI_SEULEMENT"],
                values["RCI_HORS_PERIODE"],
                values["HORS_SCOPE_RCI"],
                round(values["amount_erp"], 2),
                round(values["amount_rci"], 2),
                round(values["amount_pdf"], 2),
                round(values["amount_gap"], 2),
                round(values["montant_impacte"], 2),
            ]
        )
    return rows


def _dict_rows(records: list[dict[str, Any]], columns: list[str]) -> list[list[Any]]:
    rows = [columns]
    for record in records:
        rows.append([_clean_cell_value(record.get(column)) for column in columns])
    return rows


def _write_rows_sheet(
    workbook: Any, sheet_name: str, rows: list[list[Any]], table_name: str
) -> Any:
    worksheet = workbook.create_sheet(title=sheet_name)
    for row in rows:
        worksheet.append([_clean_cell_value(value) for value in row])

    _style_sheet(worksheet, table_name)
    return worksheet


def _style_sheet(worksheet: Any, table_name: str) -> None:
    worksheet.freeze_panes = "A2"
    if worksheet.max_row == 0 or worksheet.max_column == 0:
        return

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(bottom=Side(style="thin", color="D9E2F3"))

    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    ref = f"A1:{get_column_letter(worksheet.max_column)}{max(worksheet.max_row, 1)}"
    worksheet.auto_filter.ref = ref
    if worksheet.max_row <= MAX_TABLE_ROWS:
        table = Table(displayName=_safe_table_name(table_name), ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name=TABLE_STYLE,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    _format_special_sheet(worksheet)
    _auto_fit_columns(worksheet)


def _format_special_sheet(worksheet: Any) -> None:
    if worksheet.title != "Synthèse":
        if worksheet.title == "Synthèse par concessionnaire":
            for column_index in range(12, 17):
                for cell in worksheet.iter_cols(
                    min_col=column_index,
                    max_col=column_index,
                    min_row=2,
                    max_row=worksheet.max_row,
                ):
                    for item in cell:
                        item.number_format = '#,##0.00'
        elif worksheet.title == "RCI PDF hors période":
            _format_columns_by_index(worksheet, {5, 6, 10}, '#,##0.00')
        elif worksheet.title == "Qualité référentiel RCI":
            _format_columns_by_index(worksheet, {3}, '#,##0.00')
        return

    for row_index in range(2, worksheet.max_row + 1):
        label = worksheet.cell(row=row_index, column=1).value
        value_cell = worksheet.cell(row=row_index, column=2)
        if label in {"Résultat du contrôle batch ERP vs batch RCI", "Informations complémentaires"}:
            worksheet.cell(row=row_index, column=1).font = Font(bold=True, color="1F4E78")
            continue
        if label in SUMMARY_MONEY_LABELS:
            value_cell.number_format = '#,##0.00'
        elif label in SUMMARY_PERCENT_LABELS:
            value_cell.number_format = '0.00%'
        elif isinstance(value_cell.value, int):
            value_cell.number_format = '#,##0'
        worksheet.cell(row=row_index, column=1).font = Font(bold=True)


def _format_columns_by_index(worksheet: Any, columns: set[int], number_format: str) -> None:
    for column_index in columns:
        if column_index > worksheet.max_column:
            continue
        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row_index, column=column_index).number_format = number_format


def _format_money_columns(worksheet: Any, columns: list[str]) -> None:
    for index, column in enumerate(columns, start=1):
        if column not in MONEY_COLUMNS:
            continue
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=index).number_format = '#,##0.00'


def _apply_status_style(worksheet: Any, row_index: int, status_column: int) -> None:
    status = worksheet.cell(row=row_index, column=status_column).value
    if not status:
        return
    fill_color = STATUS_FILLS.get(str(status))
    font_color = STATUS_FONT_COLORS.get(str(status), "000000")
    if fill_color is None:
        return

    status_cell = worksheet.cell(row=row_index, column=status_column)
    status_cell.fill = PatternFill("solid", fgColor=fill_color)
    status_cell.font = Font(color=font_color, bold=True)


def _apply_severity_style(worksheet: Any, row_index: int, severity_column: int) -> None:
    severity = worksheet.cell(row=row_index, column=severity_column).value
    if not severity:
        return
    fill_color = SEVERITY_FILLS.get(str(severity))
    font_color = SEVERITY_FONT_COLORS.get(str(severity), "000000")
    if fill_color is None:
        return

    severity_cell = worksheet.cell(row=row_index, column=severity_column)
    severity_cell.fill = PatternFill("solid", fgColor=fill_color)
    severity_cell.font = Font(color=font_color, bold=True)


def _auto_fit_columns(worksheet: Any) -> None:
    max_row = min(worksheet.max_row, MAX_AUTOFIT_ROWS)
    for column_index in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        max_length = 0
        for row_index in range(1, max_row + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 55)


def _filter_status(records: list[dict[str, Any]], statuses: set[str]) -> list[dict[str, Any]]:
    return [record for record in records if record.get("status") in statuses]


def _status_counts(records: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for record in records:
        status = str(record.get("status") or "")
        if not status:
            continue
        counts[status] = counts.get(status, 0) + 1
    return counts


def _sum_amounts(records: list[dict[str, Any]], field: str) -> float:
    total = sum(abs(_number(record.get(field)) or 0) for record in records)
    return round(total, 2)


def _sum_abs_amounts(records: list[dict[str, Any]], field: str) -> float:
    total = sum(abs(_number(record.get(field)) or 0) for record in records)
    return round(total, 2)


def _number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(number) or math.isinf(number):
        return None
    return number


def _report_datetime(report: dict[str, Any], run_id: str) -> datetime:
    for date_format in ("%Y%m%d_%H%M%S", "%Y%m%d_%H%M"):
        try:
            return datetime.strptime(run_id, date_format)
        except ValueError:
            continue

    generated_at = report.get("generated_at")
    if generated_at:
        try:
            return datetime.fromisoformat(str(generated_at).replace("Z", "+00:00")).replace(tzinfo=None)
        except ValueError:
            pass
    return datetime.now()


def _clean_cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    return value


def _safe_table_name(value: str) -> str:
    cleaned = "".join(character for character in value if character.isalnum())
    if not cleaned:
        return "TableRapport"
    if cleaned[0].isdigit():
        return f"Table{cleaned}"
    return cleaned[:254]


def _yes_no(value: Any) -> str:
    return "Oui" if bool(value) else "Non"


def _require_openpyxl() -> Any:
    if Workbook is None:
        raise RuntimeError(
            "Le module openpyxl est requis pour generer le rapport Excel. "
            "Installez les dependances avec: python -m pip install -r requirements.txt"
        )
    return Workbook
